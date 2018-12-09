import h5py
import numpy as np
from openpyxl import load_workbook
from .scan_elements import Block, ScanEvent
from .read_fmri_data_abstract import FmriReader
import spacy

# This method reads the data that I received from Jonas Kaplan.
# It is described in Dehghani et al. (2017)
# Paper: https://onlinelibrary.wiley.com/doi/epdf/10.1002/hbm.23814
# They plan to publish the data eventually.

# It consists of fMRI data from 90 subjects (30 from three different native languages each: English, Mandarin, Farsi)
#  who read 40 short personal stories (145-155 words long).
# The data is already preprocessed, see section "fMRI Data Preprocessing" in the paper.
# Format: a matrix of 30 x 40 x 212018, which corresponds to subjects X stories X voxels.
# Note: the data for subject 30 is empty for English! Don't know why.
# Language should be english, chinese, or farsi

class StoryDataReader(FmriReader):
    def __init__(self, data_dir):
        super(StoryDataReader, self).__init__(data_dir)

    def read_all_events(self, subject_ids=None, **kwargs):

        self.language = kwargs.get("language", "english")
        datafile = self.data_dir + "30_" + self.language + "_storydata_masked.hd5"


        # Read stimuli and data
        data = h5py.File(datafile, 'r')
        datamatrix = np.array(data["__unnamed__"][()])

        stimulifile = self.data_dir + '/StoryKey.xlsx'
        stimuli = load_workbook(stimulifile).active

        # Set subject ids
        if subject_ids == None:
            subject_ids = list(range(0, datamatrix.shape[0]))

        # First collect all stories
        stories = []

        # We use spacy for tokenization because it is used by allennlp and thus goes well with the Elmo encoder.
        #  Something else might also work.

        tok_model = spacy.load('en_core_web_sm')
        for story_id in range(2, stimuli.max_row + 1):
            # The first 7 columns contain irrelevant information
            context = stimuli.cell(row=story_id, column=8).value.strip()
            seg1 = stimuli.cell(row=story_id, column=9).value.strip()
            seg2 = stimuli.cell(row=story_id, column=10).value.strip()
            seg3 = stimuli.cell(row=story_id, column=11).value.strip()
            story = seg1 + " " + seg2 + " " + seg3

            # I noticed some double spaces.
            story = story.replace("  ", " ")

            # Split the story into sentences
            sentences = [context.split(" ")] + self.segment_sentences(story, tok_model)
            stories.append(sentences)

        blocks = {}

        for subject in subject_ids:
            # We exclude subject 29 because the voxel activations are all 0.
            if self.language =="english" and subject ==29:
                pass
            else:
                blocks_for_subject = []
                for block_index in range(0, datamatrix.shape[1]):
                    stimulus_pointer = []
                    for sentence_id in range(0,len(stories[block_index])):
                        for word_id in range(0,len(stories[block_index][sentence_id])):
                            stimulus_pointer.append((sentence_id,word_id))

                    # For this dataset, the brain activation has already been averaged over the whole story which consists of several sentences.
                    # I do not yet have a strong opinion on whether it makes sense to include the context primer to the stimulus.

                    event = ScanEvent( str(subject),  stimulus_pointer, block_index, datamatrix[subject][block_index])
                    block = Block(str(subject), block_index, stories[block_index],[event])
                    block.scan_events = [event]
                    blocks_for_subject.append(block)
                blocks[subject] = blocks_for_subject
        return blocks


    def segment_sentences(self, story, model):
        processed = model(story)
        tokenized_sentences = []
        for sentence in processed.sents:
            tokenized_sentences.append([tok.text for tok in sentence])

        return tokenized_sentences

# def get_voxel_to_region_mapping(mapperfile, data):
# This needs to be done with pymvpa2, which is really annoying to install if you are not on Ubuntu.
# data = h5load(datafile)
# mapper = h5load(mapperfile)
# coordinates = mapper.reverse(data)
